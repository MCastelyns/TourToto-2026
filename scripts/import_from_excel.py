"""
One-time (re-run when picks change) import of participant team sheets from the
source 'tourtoto 2026 Uitslagen.xlsx' workbook into data/participants/<name>.json.

Each participant sheet has an inconsistent column offset (some start at column A,
some at column B), so this script locates the 'NAAM' header cells dynamically
instead of assuming fixed columns.
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = ROOT / "source_files" / "tourtoto 2026 Uitslagen.xlsx"
OUT_DIR = ROOT / "data" / "participants"


def find_cell(ws, text, min_row=1, max_row=60, min_col=1, max_col=10):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().upper() == text.upper():
                return cell.row, cell.column
    return None


def find_cell_contains(ws, text, min_row=1, max_row=60, min_col=1, max_col=10):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and text.upper() in cell.value.strip().upper():
                return cell.row, cell.column
    return None


def parse_participant(ws, name):
    naam_hits = []
    r = 1
    while True:
        hit = find_cell(ws, "NAAM", min_row=r, max_row=60)
        if not hit:
            break
        naam_hits.append(hit)
        r = hit[0] + 1
        if len(naam_hits) >= 2:
            break

    if len(naam_hits) < 2:
        raise ValueError(f"Could not find both NAAM blocks for {name}")

    (hoofd_header_row, naam_col), (pannen_header_row, naam_col2) = naam_hits
    rank_col = naam_col - 1
    team_col = naam_col + 1

    hoofdploeg = []
    for i in range(10):
        row = hoofd_header_row + 1 + i
        rider = ws.cell(row, naam_col).value
        team = ws.cell(row, team_col).value
        if rider:
            hoofdploeg.append({"slot": i + 1, "rider": str(rider).strip(), "team": str(team).strip() if team else None})

    joker_hit = find_cell_contains(ws, "JOKERRIT", min_row=hoofd_header_row, max_row=hoofd_header_row + 20)
    joker_stage = None
    if joker_hit:
        jrow, jcol = joker_hit
        val = ws.cell(jrow, jcol + 1).value
        if isinstance(val, (int, float)):
            joker_stage = int(val)

    rank_col2 = naam_col2 - 1
    pannenkoeken = []
    for i in range(15):
        row = pannen_header_row + 1 + i
        rider = ws.cell(row, naam_col2).value
        if rider:
            pannenkoeken.append({"slot": i + 1, "rider": str(rider).strip()})

    vragen_hit = find_cell_contains(ws, "vragen", min_row=pannen_header_row, max_row=pannen_header_row + 20)
    bonus_answers = {"podium": None, "gap_seconds": None, "dnf_count": None}
    if vragen_hit:
        vrow, vcol = vragen_hit
        podium_val = ws.cell(vrow + 1, vcol + 1).value
        gap_val = ws.cell(vrow + 2, vcol + 1).value
        dnf_val = ws.cell(vrow + 3, vcol + 1).value
        bonus_answers["podium"] = (
            [p.strip() for p in re.split(r",", str(podium_val)) if p.strip()] if podium_val else None
        )
        bonus_answers["gap_seconds"] = gap_val if isinstance(gap_val, (int, float)) else None
        bonus_answers["dnf_count"] = dnf_val if isinstance(dnf_val, (int, float)) else None

    return {
        "name": name,
        "hoofdploeg": hoofdploeg,
        "joker_stage": joker_stage,
        "pannenkoeken": pannenkoeken,
        "bonus_answers": bonus_answers,
    }


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    count = 0
    for sheet_name in wb.sheetnames:
        if sheet_name == "Totaal":
            continue
        ws = wb[sheet_name]
        data = parse_participant(ws, sheet_name)
        out_path = OUT_DIR / f"{sheet_name.lower()}.json"
        out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        n_hoofd = len(data["hoofdploeg"])
        n_pannen = len(data["pannenkoeken"])
        print(f"{sheet_name:10s} -> {out_path.name}  hoofd={n_hoofd}/10 pannen={n_pannen}/15 joker={data['joker_stage']}")
        count += 1
    print(f"\nImported {count} participants into {OUT_DIR}")


if __name__ == "__main__":
    main()
